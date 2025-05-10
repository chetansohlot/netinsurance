from multiprocessing import Value
from urllib import request
from django.http import HttpResponse
from django.contrib.auth.hashers import make_password
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render,redirect, get_object_or_404
from django.contrib import messages
from django.template import loader
from ..models import Commission, LeadUploadExcel, SourceMaster,Users, DocumentUpload, Branch, Leads, QuotationCustomer
from empPortal.model import BankDetails
from ..forms import DocumentUploadForm
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.utils.timezone import now
from django.contrib.auth import authenticate, login ,logout
from django.core.files.storage import FileSystemStorage
import re
import requests
from fastapi import FastAPI, File, UploadFile
import fitz
import openai
import time
import json
from django.http import JsonResponse
import os
import zipfile
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.db import connection
import logging
logger = logging.getLogger(__name__)
import os
import pdfkit
from django.template.loader import render_to_string
from pprint import pprint 
from django.core.paginator import Paginator
from django.db.models import Q
from empPortal.model import Referral, Partner

import pandas as pd
from django.core.files.storage import default_storage
import openpyxl
from django.db.models import Max
import re,logging
from dateutil import parser
logger = logging.getLogger(__name__)
OPENAI_API_KEY = settings.OPENAI_API_KEY
from django_q.tasks import async_task
from ..models import Users, LeadUploadExcel
from datetime import datetime, timedelta
from django.db.models import F, Value, CharField
from django.db.models.functions import Concat, Coalesce
from ..models import PolicyInfo
from ..models import Users, Roles,PolicyDocument
from django.views.decorators.csrf import csrf_exempt
from openpyxl.utils import get_column_letter
import openpyxl
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from ..model import State, City
app = FastAPI()
from datetime import datetime
from ..model import InsuranceType, InsuranceCategory, InsuranceProduct

def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def index(request):
    if not request.user.is_authenticated:
        return redirect('login')

    per_page = request.GET.get('per_page', 10)
    search_field = request.GET.get('search_field', '')
    search_query = request.GET.get('search_query', '')
    global_search = request.GET.get('global_search', '').strip()
    shorting = request.GET.get('shorting', '')  # Get sorting preference

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    # Base queryset
    if request.user.role_id != 1:
        leads = Leads.objects.filter(created_by=request.user.id)
    else:
        leads = Leads.objects.all()

    # Global search
    if global_search:
        leads = leads.filter(
            Q(name_as_per_pan__icontains=global_search) |
            Q(email_address__icontains=global_search) |
            Q(mobile_number__icontains=global_search) |
            Q(pan_card_number__icontains=global_search) |
            Q(state__icontains=global_search) |
            Q(city__icontains=global_search) |
            Q(lead_id__icontains=global_search)
        )

    # Field-specific search
    if search_field and search_query:
        filter_args = {f"{search_field}__icontains": search_query}
        leads = leads.filter(**filter_args)

     # Apply filters based on form input
    """if 'lead_id' in request.GET and request.GET['lead_id']:
        leads = leads.filter(lead_id__icontains=request.GET['lead_id'])
    if 'name_as_per_pan' in request.GET and request.GET['name_as_per_pan']:
        leads = leads.filter(name_as_per_pan__icontains=request.GET['name_as_per_pan'])    
    if 'pan_card_number' in request.GET and request.GET['pan_card_number']:
        leads = leads.filter(pan_card_number__icontains=request.GET['pan_card_number'])
    if 'email_address' in request.GET and request.GET['email_address']:
        leads = leads.filter(email_address__icontains=request.GET['email_address'])
    if 'mobile_number' in request.GET and request.GET['mobile_number']:
        leads = leads.filter(mobile_number__icontains=request.GET['mobile_number']) """
    
    # Get filter inputs
    lead_id = request.GET.get('lead_id', '')
    name = request.GET.get('name_as_per_pan', '')
    pan = request.GET.get('pan_card_number', '')
    email = request.GET.get('email_address', '')
    mobile = request.GET.get('mobile_number', '')
    sales_manager = request.GET.get('sales_manager', '')
    agent_name = request.GET.get('agent_name', '')
    policy_number = request.GET.get('policy_number', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    insurance_company = request.GET.get('insurance_company', '')
    policy_type = request.GET.get('policy_type', '')
    vehicle_type = request.GET.get('vehicle_type', '')
    upcoming_renewals = request.GET.get('upcoming_renewals', '')
    lead_type = request.GET.get('lead_type')
    motor_type = request.GET.get('motor_type')

    # Collect filter inputs
    filters_applied = any([
        lead_id, name, pan, email, mobile,
        sales_manager, agent_name, policy_number,
        start_date, end_date, insurance_company,
        policy_type, vehicle_type, upcoming_renewals,
        lead_type, motor_type
    ])

    #today = datetime.today().date()
    #after_30_days = today + timedelta(days=30)
    allowed_roles = Roles.objects.filter(roleDepartment=1)

    # Apply filters
    if lead_id:
        leads = leads.filter(lead_id__icontains=lead_id)
    if name:
        leads = leads.filter(name_as_per_pan__icontains=name)
    if pan:
        leads = leads.filter(pan_card_number__icontains=pan)
    if email:
        leads = leads.filter(email_address__icontains=email)
    if mobile:
        leads = leads.filter(mobile_number__icontains=mobile)
    if sales_manager:
        leads = leads.filter(sales_manager__user_name=sales_manager)
    if agent_name:
        leads = leads.filter(agent_name=agent_name)
    if policy_number:
        leads = leads.filter(registration_number__icontains=policy_number)
    if start_date:
        leads = leads.filter(created_at__gte=start_date)
    if end_date:
        leads = leads.filter(created_at__lte=end_date)
    if insurance_company:
        leads = leads.filter(insurance_company=insurance_company)
    if policy_type:
        leads = leads.filter(policy_type=policy_type)
    if vehicle_type:
        leads = leads.filter(vehicle_type=vehicle_type)
    if lead_type:
        leads = leads.filter(lead_type=lead_type)

    if lead_type == 'MOTOR' and motor_type:
        leads = leads.filter(vehicle_type=motor_type)

     # Example logic for upcoming renewals (next 30 days)
    upcoming_renewals = request.GET.get('upcoming_renewals')

    if upcoming_renewals:
        try:
            today = datetime.today().date()
            days = int(upcoming_renewals)
            target_date = today + timedelta(days=days)

        # Range from today to target_date
            leads = leads.filter(risk_start_date__range=[today, target_date])
        except ValueError:
            pass  # Invalid number of days (safe fallback)
    
   
    # Get unique dropdown values
    #sales_managers = Users.objects.filter(role_id=3).values('first_name','first_name', 'last_name').distinct()
    sales_managers = Users.objects.filter(
        role_id=3,
        role__in=allowed_roles
        ).values('first_name', 'last_name').distinct() 
    agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
    insurance_companies = Leads.objects.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
    policy_types = Leads.objects.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
    vehicle_types = Leads.objects.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')
   

    # Sorting
    if shorting == 'name_asc':
        leads = leads.order_by('name_as_per_pan')
    elif shorting == 'name_desc':
        leads = leads.order_by('-name_as_per_pan')
    elif shorting == 'recently_added':
        leads = leads.order_by('-created_at')
    elif shorting == 'recently_updated':
        leads = leads.order_by('-updated_at')
    else:
        leads = leads.order_by('-created_at')  # Default sort

    # After filtering leads
    #if request.GET.get('export') == '1':
        #return export_leads_to_excel(leads)
    # Handle Export
    if request.GET.get('export') == '1':
        if filters_applied:
            return export_leads_to_excel(leads)
        else:
            messages.warning(request, "Please select at least one filter to export data.")
            return redirect('leads-mgt')
    
    # Count
    all_leads = Leads.objects.all()
    total_leads = all_leads.count()  
    motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
    health_leads = Leads.objects.filter(lead_type='HEALTH').count()
    term_leads = Leads.objects.filter(lead_type='TERM').count()

    # Base queryset
    if request.user.role_id != 1:
        leads = Leads.objects.filter(created_by=request.user.id)
        all_leads = Leads.objects.filter(created_by=request.user.id)
        total_leads = all_leads.filter(created_by=request.user.id).count()  
        motor_leads = Leads.objects.filter(created_by=request.user.id,lead_type='MOTOR').count()
        health_leads = Leads.objects.filter(created_by=request.user.id,lead_type='HEALTH').count()
        term_leads = Leads.objects.filter(created_by=request.user.id,lead_type='TERM').count()
    else:
        all_leads = Leads.objects.all()
        total_leads = all_leads.count()  
        motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
        health_leads = Leads.objects.filter(lead_type='HEALTH').count()
        term_leads = Leads.objects.filter(lead_type='TERM').count()

    # Pagination
    paginator = Paginator(leads, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'leads/index.html', {
        'page_obj': page_obj,
        'total_leads': total_leads,
        'per_page': per_page,
        'search_field': search_field,
        'search_query': search_query,
        'shorting': shorting,  # Pass to template to retain selected option
        'motor_leads': motor_leads,
        'health_leads': health_leads,
        'term_leads': term_leads,
        'sales_managers': sales_managers,
        'selected_sales_manager': sales_manager,
        'agents': agents,
        'selected_agent': agent_name,
        'insurance_companies': insurance_companies,
        'policy_types': policy_types,
        'vehicle_types': vehicle_types,
        'leads': leads,
    })

def export_leads_to_excel(leads_queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Step 1: Check if any lead has a policy
    has_policy_data = False
    for lead in leads_queryset:
        policy = PolicyDocument.objects.filter(vehicle_number=lead.registration_number).order_by('-created_at').first()
        if policy and policy.policy_number:
            has_policy_data = True
            break

    # Step 2: Define headers dynamically
    headers = [
        'S.No', 'Lead ID', 'Name as per PAN', 'Email Address', 'Mobile Number',
        'PAN Card Number', 'Vehicle Number'
    ]
    if has_policy_data:
        headers += [
            'Previous Policy Number', 'Policy Issue Date', 'Policy Expiry Date',
            'Sum Insured', 'Net Insurance', 'Gross Insurance'
        ]
    headers += ['Created Date']

    ws.append(headers)
    ws.freeze_panes = 'A2'

    # Step 3: Write rows
    for index, lead in enumerate(leads_queryset, start=1):
        previous_policy = PolicyDocument.objects.filter(vehicle_number=lead.registration_number).order_by('-created_at').first()

        row = [
            index,
            lead.id,
            lead.name_as_per_pan or '',
            lead.email_address or '',
            lead.mobile_number or '',
            lead.pan_card_number or '',
            lead.registration_number or '',
          
        ]

        if has_policy_data:
            row += [
                previous_policy.policy_number if previous_policy else '',
                previous_policy.policy_issue_date if previous_policy else '',
                previous_policy.policy_expiry_date if previous_policy else '',
                previous_policy.sum_insured if previous_policy else '',
                previous_policy.policy_premium if previous_policy else '',
                previous_policy.policy_total_premium if previous_policy else '',
            ]

        row += [
            lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else '',
            
        ]
            
        ws.append(row)

    # Step 4: Adjust column widths
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = max_length + 2

    # Step 5: Return Excel response
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="filtered_leads.xlsx"'

    return response


def viewlead(request, lead_id):
    if request.user.is_authenticated:
        leads = Leads.objects.all()
        return render(request, 'leads/index.html', {
            'leads': leads,})  # Pass leads to the template
    else:
        return redirect('login')
    
def healthLead(request):
    if request.user.is_authenticated:

        leads = Leads.objects.all()
        # Count
        all_leads = Leads.objects.all()
        total_leads = all_leads.count()  
        motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
        health_leads = Leads.objects.filter(lead_type='HEALTH').count()
        term_leads = Leads.objects.filter(lead_type='TERM').count()

        # Get filter inputs
        lead_id = request.GET.get('lead_id', '')
        name = request.GET.get('name_as_per_pan', '')
        pan = request.GET.get('pan_card_number', '')
        email = request.GET.get('email_address', '')
        mobile = request.GET.get('mobile_number', '')
        sales_manager = request.GET.get('sales_manager', '')
        agent_name = request.GET.get('agent_name', '')
        policy_number = request.GET.get('policy_number', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        insurance_company = request.GET.get('insurance_company', '')
        policy_type = request.GET.get('policy_type', '')
        vehicle_type = request.GET.get('vehicle_type', '')
        upcoming_renewals = request.GET.get('upcoming_renewals', '')
        lead_type = request.GET.get('lead_type')
        motor_type = request.GET.get('motor_type')

    #today = datetime.today().date()
    #after_30_days = today + timedelta(days=30)
    # Apply filters
        if lead_id:
            leads = leads.filter(lead_id__icontains=lead_id)
        if name:
            leads = leads.filter(name_as_per_pan__icontains=name)
        if pan:
            leads = leads.filter(pan_card_number__icontains=pan)
        if email:
            leads = leads.filter(email_address__icontains=email)
        if mobile:
            leads = leads.filter(mobile_number__icontains=mobile)
        if sales_manager:
            leads = leads.filter(sales_manager__user_name=sales_manager)
        if agent_name:
            leads = leads.filter(agent_name=agent_name)
        if policy_number:
            leads = leads.filter(registration_number__icontains=policy_number)
        if start_date:
            leads = leads.filter(risk_start_date__gte=start_date)
        if end_date:
            leads = leads.filter(risk_start_date__lte=end_date)
        if insurance_company:
            leads = leads.filter(insurance_company=insurance_company)
        if policy_type:
            leads = leads.filter(policy_type=policy_type)
        if vehicle_type:
            leads = leads.filter(vehicle_type=vehicle_type)
        if lead_type:
            leads = leads.filter(lead_type=lead_type)

        if lead_type == 'MOTOR' and motor_type:
            leads = leads.filter(vehicle_type=motor_type)

     # Example logic for upcoming renewals (next 30 days)
        upcoming_renewals = request.GET.get('upcoming_renewals')

        if upcoming_renewals:
            try:
                today = datetime.today().date()
                days = int(upcoming_renewals)
                target_date = today + timedelta(days=days)

        # Range from today to target_date
                leads = leads.filter(risk_start_date__range=[today, target_date])
            except ValueError:
                pass  # Invalid number of days (safe fallback)
    
   
        # Get unique dropdown values
        sales_managers = Users.objects.filter(role_id=3).values('first_name','first_name', 'last_name').distinct()
      
        agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
        insurance_companies = Leads.objects.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
        policy_types = Leads.objects.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
        vehicle_types = Leads.objects.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')

         # After filtering leads
        if request.GET.get('export') == '1':
            return export_leads_to_excel(leads)

        return render(request, 'leads/health-lead.html',{
            'leads': leads,
            'total_leads':total_leads,
            'motor_leads': motor_leads,
            'health_leads': health_leads,
            'term_leads': term_leads,
            'sales_managers': sales_managers,
            'selected_sales_manager': sales_manager,
            'agents': agents,
            'selected_agent': agent_name,
            'insurance_companies': insurance_companies,
            'policy_types': policy_types,
            #'vehicle_types': vehicle_types,
            
        })  
    else:
        return redirect('login')
    
def termlead(request):
    if request.user.is_authenticated:

        leads = Leads.objects.all()
        # Count
        all_leads = Leads.objects.all()
        total_leads = all_leads.count()  
        motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
        health_leads = Leads.objects.filter(lead_type='HEALTH').count()
        term_leads = Leads.objects.filter(lead_type='TERM').count()

        # Get filter inputs
        lead_id = request.GET.get('lead_id', '')
        name = request.GET.get('name_as_per_pan', '')
        pan = request.GET.get('pan_card_number', '')
        email = request.GET.get('email_address', '')
        mobile = request.GET.get('mobile_number', '')
        sales_manager = request.GET.get('sales_manager', '')
        agent_name = request.GET.get('agent_name', '')
        policy_number = request.GET.get('policy_number', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        insurance_company = request.GET.get('insurance_company', '')
        policy_type = request.GET.get('policy_type', '')
        vehicle_type = request.GET.get('vehicle_type', '')
        upcoming_renewals = request.GET.get('upcoming_renewals', '')
        lead_type = request.GET.get('lead_type')
        motor_type = request.GET.get('motor_type')

    #today = datetime.today().date()
    #after_30_days = today + timedelta(days=30)
    # Apply filters
        if lead_id:
            leads = leads.filter(lead_id__icontains=lead_id)
        if name:
            leads = leads.filter(name_as_per_pan__icontains=name)
        if pan:
            leads = leads.filter(pan_card_number__icontains=pan)
        if email:
            leads = leads.filter(email_address__icontains=email)
        if mobile:
            leads = leads.filter(mobile_number__icontains=mobile)
        if sales_manager:
            leads = leads.filter(sales_manager__user_name=sales_manager)
        if agent_name:
            leads = leads.filter(agent_name=agent_name)
        if policy_number:
            leads = leads.filter(registration_number__icontains=policy_number)
        if start_date:
            leads = leads.filter(risk_start_date__gte=start_date)
        if end_date:
            leads = leads.filter(risk_start_date__lte=end_date)
        if insurance_company:
            leads = leads.filter(insurance_company=insurance_company)
        if policy_type:
            leads = leads.filter(policy_type=policy_type)
        if vehicle_type:
            leads = leads.filter(vehicle_type=vehicle_type)
        if lead_type:
            leads = leads.filter(lead_type=lead_type)

        if lead_type == 'MOTOR' and motor_type:
            leads = leads.filter(vehicle_type=motor_type)

     # Example logic for upcoming renewals (next 30 days)
        upcoming_renewals = request.GET.get('upcoming_renewals')

        if upcoming_renewals:
            try:
                today = datetime.today().date()
                days = int(upcoming_renewals)
                target_date = today + timedelta(days=days)

        # Range from today to target_date
                leads = leads.filter(risk_start_date__range=[today, target_date])
            except ValueError:
                pass  # Invalid number of days (safe fallback)
    
   
        # Get unique dropdown values
        sales_managers = Users.objects.filter(role_id=3).values('first_name','first_name', 'last_name').distinct()
        agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
        insurance_companies = Leads.objects.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
        policy_types = Leads.objects.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
        vehicle_types = Leads.objects.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')
   


        return render(request, 'leads/term-lead.html',
        {
            'leads': leads,
            'total_leads':total_leads,
            'motor_leads': motor_leads,
            'health_leads': health_leads,
            'term_leads': term_leads,
            'sales_managers': sales_managers,
            'selected_sales_manager': sales_manager,
            'agents': agents,
            'selected_agent': agent_name,
            'insurance_companies': insurance_companies,
            'policy_types': policy_types,
            'vehicle_types': vehicle_types,
        })  # Pass leads to the template
    else:
        return redirect('login')

def lead_init_view(request):
    types = InsuranceType.objects.all()
    return render(request, 'leads/lead-init.html', {'types': types})

# For AJAX - Load categories
def load_categories(request):
    type_id = request.GET.get('insurance_type')
    categories = InsuranceCategory.objects.filter(insurance_type_id=type_id).values('id', 'name')
    return JsonResponse(list(categories), safe=False)

# For AJAX - Load products
def load_products(request):
    category_id = request.GET.get('insurance_category')
    products = InsuranceProduct.objects.filter(category_id=category_id).values('id', 'name')
    return JsonResponse(list(products), safe=False)



def create_or_edit_lead(request, lead_id=None):
    if not request.user.is_authenticated:
        return redirect('login')
    
    customers = QuotationCustomer.objects.all()
    source_leads = SourceMaster.objects.filter(status=True).order_by('source_name')
    print("Loaded sources:", list(source_leads))

    lead = None
    referrals = Referral.objects.all()
    partners = Partner.objects.filter(active=True)

    if lead_id:
        lead = get_object_or_404(Leads, id=lead_id)
    else:
        lead = None  
    
    states = State.objects.all()
    
    
    if request.method == "GET":
        return render(request, 'leads/create.html', {
            'lead': lead,
            'referrals': referrals,
            'customers': customers,
            'states': states,
            'partners':partners
        })


    elif request.method == "POST":
        mobile_number = request.POST.get("mobile_number", "").strip()
        email_address = request.POST.get("email_address", "").strip()
        quote_date = request.POST.get("quote_date", None)
        name_as_per_pan = request.POST.get("name_as_per_pan", "").strip()
        pan_card_number = request.POST.get("pan_card_number", "").strip() or None


        source_leads_id = request.POST.get("source_leads", "").strip() or None
        source_leads = SourceMaster.objects.get(id=source_leads_id) if source_leads_id else None
        
        # ✅ Handle date_of_birth safely
        date_of_birth_str = request.POST.get("date_of_birth", "").strip()
        date_of_birth = None
        if date_of_birth_str:
            try:
                date_of_birth = datetime.strptime(date_of_birth_str, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format for Date of Birth. Please use YYYY-MM-DD.")
                return redirect(request.path)

        #state = request.POST.get("state", "").strip()
        #city = request.POST.get("city", "").strip()
        state_name = request.POST.get('state')
        city_name = request.POST.get('city')
        pincode = request.POST.get("pincode", "").strip()
        address = request.POST.get("address", "").strip()
        lead_source = request.POST.get("lead_source", "").strip()
        referral_by = request.POST.get("referral_by", "").strip()
        partner_id = request.POST.get("partner_id", "").strip()
        if lead_source != 'referral_partner':
            referral_by = ''
        if lead_source != 'pos_partner':
            partner_id = None
        lead_description = request.POST.get("lead_description", "").strip()
        # lead_type = request.POST.get("lead_type", "MOTOR").strip()
        lead_type = request.POST.get("lead_type", "").strip()
        #motor_leads = Leads.objects.filter(lead_type='MOTOR', created_by=request.user).order_by('-id')
        #health_leads = Leads.objects.filter(lead_type='HEALTH', created_by=request.user).order_by('-id')
        #term_leads = Leads.objects.filter(lead_type='TERM', created_by=request.user).order_by('-id')

        if lead_type == "MOTOR":
            registration_number = request.POST.get("registration_number", "").strip()
            
            vehicle_type = request.POST.get("vehicle_type", "").strip()

        else:
           registration_number = ""
           vehicle_type = ""  

        # Before creating Lead, try to fetch matching PolicyDocument
        policy_document = None
        if registration_number:
                policy_document = PolicyDocument.objects.filter(vehicle_number=registration_number).first()

                policy_start_date_str = policy_document.policy_start_date if policy_document else None
                try:
                    # Try parsing with both date and time
                    if policy_start_date_str:
                        policy_start_date_obj = datetime.strptime(policy_start_date_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        # If there's no time component, parse as date only
                        if policy_start_date_str:
                            policy_start_date_obj = datetime.strptime(policy_start_date_str, '%Y-%m-%d')
                    except ValueError:
                        policy_start_date_obj = None

        # Default fields
        insurance_company = None
        policy_number = None
        policy_type = None
        vehicle_type_db = None
        sum_insured = None
        policy_date = None

        # If matching policy found, extract details
        if policy_document:
            insurance_company = policy_document.insurance_provider
            policy_number = policy_document.policy_number
            policy_type = policy_document.policy_type
            vehicle_type_db = policy_document.vehicle_type
            sum_insured = policy_document.sum_insured
            policy_date = policy_start_date_obj.strftime('%Y-%m-%d %H:%M:%S')
           
       
        status = request.POST.get("status", "new").strip()
        if lead:
            lead.mobile_number = mobile_number
            lead.email_address = email_address
            lead.quote_date = quote_date
            lead.name_as_per_pan = name_as_per_pan
            lead.pan_card_number = pan_card_number
            lead.date_of_birth = date_of_birth
            lead.state = state_name
            lead.city = city_name
            lead.pincode = pincode
            lead.address = address
            lead.lead_description = lead_description
            lead.lead_type = lead_type
            lead.registration_number = registration_number  ## vehicle no.
            lead.vehicle_type = vehicle_type
            lead.lead_source = lead_source
            lead.referral_by = referral_by
            lead.partner_id = partner_id
            lead.status = status
            lead.updated_at = now()
            lead.insurance_company=insurance_company
            lead.policy_number=policy_number
            lead.policy_type=policy_type
            lead.sum_insured=sum_insured
            lead.risk_start_date=policy_date
            lead.save()
            messages.success(request, f"Lead updated successfully! Lead ID: {lead.lead_id}")
        else:
            new_lead = Leads.objects.create(
                mobile_number=mobile_number,
                email_address=email_address,
                quote_date=quote_date,
                name_as_per_pan=name_as_per_pan,
                pan_card_number=pan_card_number,
                date_of_birth=date_of_birth,
                state=state_name,
                city=city_name,
                pincode=pincode,
                address=address,
                lead_description=lead_description,
                lead_type=lead_type,
                registration_number=registration_number,
                vehicle_type=vehicle_type,
                #vehicle_type = request.POST.get("vehicle_type", None)
                lead_source=lead_source,
                referral_by=referral_by,
                status=status,
                created_by=request.user.id,
                created_at=now(),
                updated_at=now(),
                insurance_company=insurance_company,
                policy_number=policy_number,
                policy_type=policy_type,
                sum_insured=sum_insured,
                risk_start_date = policy_date,
                partner_id = partner_id
                
            )

            # Step 2: Generate lead_id using date + ID
           
            today_str = datetime.today().strftime('%Y%m%d')  # e.g., 20250508
            lead_id = f"{today_str}{new_lead.id}"   

            # Step 3: Save it to the lead
            new_lead.lead_id = lead_id
            new_lead.save()
            messages.success(request, f"Lead created successfully!  {lead_id}")

        return redirect("leads-mgt")

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime

@csrf_exempt
def fetch_policy_details(request):
    if request.method == "POST":
        registration_number = request.POST.get("registration_number")
        try:
            policy_document = PolicyDocument.objects.filter(vehicle_number=registration_number).first()
            
            if policy_document:  #First check if found
                policy_start_date_str = policy_document.policy_start_date
                policy_start_date_obj = datetime.strptime(policy_start_date_str, '%Y-%m-%d %H:%M:%S') if policy_start_date_str else None

                return JsonResponse({
                    'success': True,
                    'insurance_company': policy_document.insurance_provider,
                    'policy_number': policy_document.policy_number,
                    'policy_type': policy_document.policy_type,
                    'vehicle_type': policy_document.vehicle_type,
                    'sum_insured': policy_document.sum_insured,
                    'policy_date': policy_start_date_obj.strftime('%Y-%m-%d %H:%M:%S') if policy_start_date_obj else "",
                })
            else:
                # ✅ If no policy found
                return JsonResponse({'success': False, 'message': 'Please fill in all details'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request'})
#state
def get_state(request):
    states = State.objects.all()
    return render(request, 'leads/create-location-info.html', {'states': states})

def get_cities(request):
    """state_name = request.GET.get('state_id')
    try:
        state = State.objects.get(name=state_name)
        cities = City.objects.filter(state=state).values('city')
        return JsonResponse(list(cities), safe=False)
    except State.DoesNotExist:
        return JsonResponse([], safe=False)"""
    state_id = request.GET.get('state_id')
    cities = City.objects.filter(state_id=state_id).values('id', 'city')
    return JsonResponse({'cities': list(cities)})

"""def bulk_upload_leads(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "Please select a file to upload.")
            return redirect('leads-mgt')

        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect('leads-mgt')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Get latest lead_id once before loop
            latest_lead = Leads.objects.aggregate(Max('lead_id'))['lead_id__max']
            start_num = 1
            if latest_lead:
                match = re.search(r'L(\d+)', latest_lead)
                if match:
                    start_num = int(match.group(1)) + 1

            inserted = 0
            duplicate_data_found = False

            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 13 or any(cell is None for cell in row[:13]):
                    logger.warning(f"Row {index} skipped: Incomplete data.")
                    continue

                # Extract values
                mobile_number = str(row[0]).strip()
                email_address = str(row[1]).strip()
                quote_date = row[2]
                name_as_per_pan = str(row[3]).strip()
                pan_card_number = str(row[4]).strip().upper()
                date_of_birth = row[5]
                state = str(row[6]).strip()
                city = str(row[7]).strip()
                pincode = str(row[8]).strip()
                lead_source = str(row[9]).strip()
                address = str(row[10]).strip()
                lead_description = str(row[11]).strip()
                lead_type = str(row[12]).strip()

                # VALIDATION
                if not re.fullmatch(r'[6-9]\d{9}', mobile_number):
                    logger.warning(f"Row {index} skipped: Invalid mobile number - {mobile_number}")
                    continue

                if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email_address):
                    logger.warning(f"Row {index} skipped: Invalid email - {email_address}")
                    continue

                if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', pan_card_number):
                    logger.warning(f"Row {index} skipped: Invalid PAN - {pan_card_number}")
                    continue

                try:
                    dob = parser.parse(str(date_of_birth)).date()
                except Exception as e:
                    logger.warning(f"Row {index} skipped: Invalid DOB - {date_of_birth}")
                    continue

                # Duplicate Check
                if Leads.objects.filter(
                    Q(mobile_number=mobile_number) |
                    Q(email_address=email_address) |
                    Q(pan_card_number=pan_card_number)
                ).exists():
                    logger.info(f"Row {index} skipped: Duplicate entry - Mobile: {mobile_number}, Email: {email_address}, PAN: {pan_card_number}")
                    duplicate_data_found = True
                    continue

                # Generate lead_id
                lead_id = f"L{start_num:05d}"
                start_num += 1

                # Insert into DB
                try:
                    Leads.objects.create(
                        lead_id=lead_id,
                        mobile_number=mobile_number,
                        email_address=email_address,
                        quote_date=quote_date,
                        name_as_per_pan=name_as_per_pan,
                        pan_card_number=pan_card_number,
                        date_of_birth=dob,
                        state=state,
                        city=city,
                        pincode=pincode,
                        lead_source=lead_source,
                        address=address,
                        lead_description=lead_description,
                        lead_type=lead_type,
                    )
                    inserted += 1

                except Exception as e:
                    logger.error(f"Row {index} error: {e}")
                    continue

            # Final message
            if inserted > 0:
                messages.success(request, f"{inserted} leads uploaded successfully.")
            elif duplicate_data_found:
                messages.warning(request, "No new leads were inserted. All records were duplicates.")
            else:
                messages.info(request, "No valid data found in Excel file!")

            return redirect('leads-mgt')

        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            messages.error(request, f"Error processing file: {e}")
            return redirect('leads-mgt')

    return render(request, 'leads/bulk_upload.html')"""


"""def bulk_upload_leads(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "Please select a file to upload.")
            return redirect('leads-mgt')

        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect('leads-mgt')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            latest_lead = Leads.objects.aggregate(Max('lead_id'))['lead_id__max']
            start_num = 1
            if latest_lead:
                match = re.search(r'L(\d+)', latest_lead)
                if match:
                    start_num = int(match.group(1)) + 1

            inserted = 0
            duplicate_data_found = False

            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 29 or any(cell is None for cell in row[:13]):
                    logger.warning(f"Row {index} skipped: Incomplete data.")
                    continue

                try:
                    mobile_number = str(row[0]).strip()
                    email_address = str(row[1]).strip()
                    quote_date = row[2]
                    name_as_per_pan = str(row[3]).strip()
                    pan_card_number = str(row[4]).strip().upper()
                    date_of_birth = row[5]
                    state = str(row[6]).strip()
                    city = str(row[7]).strip()
                    pincode = str(row[8]).strip()
                    lead_source = str(row[9]).strip()
                    address = str(row[10]).strip()
                    lead_description = str(row[11]).strip()
                    lead_type = str(row[12]).strip()

                    # New fields
                    policy_date = row[13]
                    sales_manager = str(row[14]).strip()
                    agent_name = str(row[15]).strip()
                    insurance_company = str(row[16]).strip()
                    policy_type = str(row[17]).strip()
                    policy_number = str(row[18]).strip()
                    vehicle_type = str(row[19]).strip()
                    make_and_model = str(row[20]).strip()
                    fuel_type = str(row[21]).strip()
                    registration_number = str(row[22]).strip()
                    manufacturing_year = row[23]
                    sum_insured = row[24]
                    ncb = row[25]
                    od_premium = row[26]
                    tp_premium = row[27]
                    risk_start_date = row[28]

                    # VALIDATIONS
                    if not re.fullmatch(r'[6-9]\d{9}', mobile_number):
                        logger.warning(f"Row {index} skipped: Invalid mobile number - {mobile_number}")
                        continue

                    if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email_address):
                        logger.warning(f"Row {index} skipped: Invalid email - {email_address}")
                        continue

                    if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', pan_card_number):
                        logger.warning(f"Row {index} skipped: Invalid PAN - {pan_card_number}")
                        continue

                    try:
                        dob = parser.parse(str(date_of_birth)).date()
                    except Exception as e:
                        logger.warning(f"Row {index} skipped: Invalid DOB - {date_of_birth}")
                        continue

                    try:
                        policy_date = parser.parse(str(policy_date)).date() if policy_date else None
                    except:
                        policy_date = None

                    try:
                        risk_start_date = parser.parse(str(risk_start_date)).date() if risk_start_date else None
                    except:
                        risk_start_date = None

                    # Duplicate check
                    if Leads.objects.filter(
                        Q(mobile_number=mobile_number) |
                        Q(email_address=email_address) |
                        Q(pan_card_number=pan_card_number)
                    ).exists():
                        logger.info(f"Row {index} skipped: Duplicate entry.")
                        duplicate_data_found = True
                        continue

                    # Create new lead_id
                    lead_id = f"L{start_num:05d}"
                    start_num += 1

                    # Insert into DB
                    Leads.objects.create(
                        lead_id=lead_id,
                        mobile_number=mobile_number,
                        email_address=email_address,
                        quote_date=quote_date,
                        name_as_per_pan=name_as_per_pan,
                        pan_card_number=pan_card_number,
                        date_of_birth=dob,
                        state=state,
                        city=city,
                        pincode=pincode,
                        lead_source=lead_source,
                        address=address,
                        lead_description=lead_description,
                        lead_type=lead_type,
                        policy_date=policy_date,
                        sales_manager=sales_manager,
                        agent_name=agent_name,
                        insurance_company=insurance_company,
                        policy_type=policy_type,
                        policy_number=policy_number,
                        vehicle_type=vehicle_type,
                        make_and_model=make_and_model,
                        fuel_type=fuel_type,
                        registration_number=registration_number,
                        manufacturing_year=manufacturing_year,
                        sum_insured=sum_insured,
                        ncb=ncb,
                        od_premium=od_premium,
                        tp_premium=tp_premium,
                        net_premium=(od_premium or 0) + (tp_premium or 0),
                        gross_premium=(od_premium or 0) + (tp_premium or 0),
                        risk_start_date=risk_start_date
                    )
                    inserted += 1

                except Exception as e:
                    logger.error(f"Row {index} error: {e}")
                    continue

            # Final messages
            if inserted > 0:
                messages.success(request, f"{inserted} leads uploaded successfully.")
            elif duplicate_data_found:
                messages.warning(request, "No new leads were inserted. All records were duplicates.")
            else:
                messages.info(request, "No valid data found in Excel file!")

            return redirect('leads-mgt')

        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            messages.error(request, f"Error processing file: {e}")
            return redirect('leads-mgt')

    return render(request, 'leads/bulk_upload.html')"""


@login_required
def bulk_upload_leads(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        camp_name = request.POST.get("camp_name")

        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, "Only Excel files (.xlsx, .xls) are allowed!")
            return redirect('bulk-upload-leads')

        if not camp_name:
            messages.error(request, "Campaign Name is mandatory.")
            return redirect('bulk-upload-leads')


        # Save the file record
        instance = LeadUploadExcel.objects.create(
            file=excel_file,
            file_name=excel_file.name,
            file_url=excel_file.name,
            campaign_name=camp_name,
            created_by=request.user  # assuming request.user is correct
        )

        # Trigger background task
        async_task("empPortal.taskz.process_lead_excel.process_lead_excel", instance.id)

        messages.success(request, "File uploaded. It will be processed in the background.")
        return redirect("leads-mgt")

    return render(request, "leads/bulk_upload.html")

#for ctraye lead step by step
def basic_info(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    return render(request, "leads/create-basic-details.html", {'lead_id': lead_id})
    
def lead_source(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    return render(request, "leads/create-lead-source-info.html",{"lead_ref_id":lead_id})

def lead_location(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    return render(request, "leads/create-location-info.html",{"lead_ref_id":lead_id})

def assignment(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    return render(request, "leads/create-assignment.html",{"lead_ref_id":lead_id})

def previous_policy_info(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    return render(request, "leads/create.html",{"lead_ref_id":lead_id})


def save_leads_insurance_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_insurance_type_id = request.POST.get('insurance_type').strip()
    lead_insurance_category_id = request.POST.get('insurance_category').strip()
    lead_insurance_product_id = request.POST.get('insurance_product').strip()
    lead_first_name = request.POST.get('first_name').strip()
    lead_last_name = request.POST.get('last_name').strip()
    mobile_number = request.POST.get('mobile').strip()
    
    try:
        leads_insert = Leads.objects.create(
            lead_id = int(time.time()),
            lead_insurance_type_id = lead_insurance_type_id,
            lead_insurance_category_id = lead_insurance_category_id,
            lead_insurance_product_id = lead_insurance_product_id,
            lead_first_name = lead_first_name,
            lead_last_name = lead_last_name,
            mobile_number = mobile_number
        )
        
        lead_ref_id = leads_insert.lead_id
        
        messages.success(request,f"Saved Succesfully")
        return redirect('basic-info',lead_id=lead_ref_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_insurance_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_basic_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_id').strip()
    customer_name = request.POST.get('customer_name').strip()
    mobile_number = request.POST.get('mobile_number').strip()
    email_address = request.POST.get('email_address').strip()
    gender = request.POST.get('gender').strip()
    date_of_birth = request.POST.get('date_of_birth').strip()
    identity_no = request.POST.get('identity_no').strip()
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.name_as_per_pan = customer_name
        lead_data.mobile_number = mobile_number
        lead_data.email_address = email_address
        lead_data.lead_customer_gender = gender
        lead_data.date_of_birth = date_of_birth
        lead_data.lead_customer_identity_no = identity_no
        lead_data.save()
        
        lead_id = lead_data.lead_id
        messages.success(request,f"Saved Succesfully")
        return redirect('lead-source',lead_id=lead_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_basic_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_source_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id').strip()
    lead_source = request.POST.get('lead_source_name').strip()
    refered_by = request.POST.get('refered_by').strip()
    referral_name = request.POST.get('referral_name').strip()
    referral_mobile_number = request.POST.get('referral_mobile_number').strip()
    source_medium = request.POST.get('source_medium').strip()
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.lead_source = lead_source
        lead_data.referral_by = refered_by
        lead_data.referral_name = referral_name
        lead_data.referral_mobile_no = referral_mobile_number
        lead_data.lead_source_medium = source_medium
        lead_data.save()
        
        lead_ref_id = lead_data.lead_id
        messages.success(request,f"Saved Succesfully")
        return redirect('lead-location',lead_id=lead_ref_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_source_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_location_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id').strip()
    state = request.POST.get('state').strip()
    city = request.POST.get('city').strip()
    pincode = request.POST.get('pincode').strip()
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.state = state
        lead_data.city = city
        lead_data.pincode = pincode
        lead_data.save()
        
        lead_ref_id = lead_data.lead_id
        messages.success(request,f"Saved Succesfully")
        return redirect('lead-assignment',lead_id=lead_ref_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_location_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_assignment_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id').strip()
    assigned_to = request.POST.get('assigned_to').strip()
    branch = request.POST.get('branch').strip()
    lead_status_type = request.POST.get('lead_status_type').strip()
    lead_tag = request.POST.get('lead_tag').strip()
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.assigned_to = assigned_to
        lead_data.branch_id = branch
        lead_data.lead_status_type = lead_status_type
        lead_data.lead_tag = lead_tag
        lead_data.save()
            
        lead_ref_id = lead_data.lead_id
        messages.success(request,f"Saved Succesfully")
        return redirect('leads-previous-policy-info',lead_id=lead_ref_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_assignment_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_previous_policy_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id').strip()
    previous_insurer_name = request.POST.get('previous_insurer_name').strip()
    policy_number = request.POST.get('policy_number').strip()
    policy_type = request.POST.get('policy_type').strip()
    policy_date = request.POST.get('policy_date').strip()
    policy_end_date = request.POST.get('policy_end_date').strip()
    expiry_status = request.POST.get('expiry_status').strip()
    ncb = request.POST.get('ncb').strip()
    previous_idv_amount = request.POST.get('previous_idv_amount').strip()
    previous_sum_insured = request.POST.get('previous_sum_insured').strip()
    claim_made = request.POST.get('claim_made').strip()
    claim_amount = request.POST.get('claim_amount').strip()
    previous_policy_source = request.POST.get('previous_policy_source').strip()
    vehicle_type = request.POST.get('vehicle_type').strip()
    vehicle_class = request.POST.get('vehicle_class').strip()
    insurance_type = request.POST.get('insurance_type').strip()
    product_category = request.POST.get('product_category').strip()
    vehicle_reg_no = request.POST.get('vehicle_reg_no').strip()
    vehicle_make = request.POST.get('vehicle_make').strip()
    vehicle_model = request.POST.get('vehicle_model').strip()
    mgf_year = request.POST.get('mgf_year').strip()
    sum_insured = request.POST.get('sum_insured').strip()
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.previous_insurer_name = previous_insurer_name
        lead_data.policy_number = policy_number
        lead_data.policy_type = policy_type
        lead_data.policy_date = policy_date
        lead_data.policy_end_date = policy_end_date
        lead_data.expiry_status = expiry_status
        lead_data.ncb = ncb
        lead_data.previous_idv_amount = previous_idv_amount
        lead_data.previous_sum_insured = previous_sum_insured
        lead_data.claim_made = claim_made
        lead_data.claim_amount = claim_amount
        lead_data.previous_policy_source = previous_policy_source
        lead_data.vehicle_type = vehicle_type
        lead_data.vehicle_class = vehicle_class
        lead_data.insurance_type = insurance_type
        lead_data.product_category = product_category
        lead_data.vehicle_reg_no = vehicle_reg_no
        lead_data.vehicle_make = vehicle_make
        lead_data.vehicle_model = vehicle_model
        lead_data.mgf_year = mgf_year
        lead_data.sum_insured = sum_insured
        lead_data.save()
            
        messages.success(request,f"Saved Succesfully")
        return redirect('leads-mgt')
        
    except Exception as e:
        logger.error(f"Error in save_leads_previous_policy_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def view_lead(request, lead_id):
    if not request.user.is_authenticated and request.user.is_active != 1:
        return redirect('login')
        
    lead = get_object_or_404(Leads, lead_id=lead_id)
    return render(request, 'leads/view-lead.html', {'lead': lead})