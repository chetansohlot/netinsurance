from django.http import HttpResponse
import pandas as pd 
from django.contrib.auth.hashers import make_password
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render,redirect
from django.contrib import messages
from django.template import loader
from .models import Roles,Users,PolicyDocument,BulkPolicyLog
from django.contrib.auth import authenticate, login ,logout
from django.core.files.storage import FileSystemStorage
import re
import requests
from fastapi import FastAPI, File, UploadFile
import fitz
import openai
import time
import json
from django.http import JsonResponse
import os
import zipfile
from django.contrib import messages
from .models import PolicyDocument
from faker import Faker 
from .models import PolicyDocument ,Commission, AgentPaymentDetails, InsurerPaymentDetails
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
fake = Faker()
from django.utils import timezone
import datetime
from django.conf import settings
from django.db.models import Q
from django.core.paginator import Paginator


dt_aware = timezone.now()  # Django returns a timezone-aware datetime
dt_naive = dt_aware.replace(tzinfo=None) 

processed_text = {"policy_number": "3005/O/379425038/00/000", "vehicle_number": "HR98P4781", "insured_name": "SHELLEY MUNJAL", "issue_date": "2025-02-01", "expiry_date": "2026-02-01", "premium_amount": "1,163.00", "sum_insured": "64,073.00", "policy_period": "1 year", "total_premium": "1,372.00", "insurance_company": "ICICI Lombard General Insurance Company Limited", "coverage_details": [{"benefit": "Basic OD Premium", "amount": "612.00"}, {"benefit": "Zero Depreciation (Silver)", "amount": "449.00"}, {"benefit": "Return to Invoice", "amount": "224.00"}]}
    
def exportPolicies(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request, "Please Login First")
        return redirect('login')
    # Query the PolicyDocument model for all policy records
    policies = PolicyDocument.objects.all().order_by('-id')

    # Prepare the data for export
    policy_data = []
    for policy in policies:
        policy_data.append([
            policy.insurance_provider,
            policy.vehicle_number,
            policy.holder_name,
            policy.policy_number,
            policy.policy_issue_date,
            policy.policy_expiry_date,
            policy.policy_period,
            policy.policy_premium,
            policy.policy_total_premium,
            policy.payment_status,  # Add payment_status
            policy.policy_type,  # Add policy_type
            policy.vehicle_type,  # Add vehicle_type
            policy.vehicle_make,  # Add vehicle_make
            policy.vehicle_model,  # Add vehicle_model
            policy.vehicle_gross_weight,  # Add vehicle_gross_weight
            policy.vehicle_manuf_date,  # Add vehicle_manuf_date
            policy.gst,  # Add gst
            policy.od_premium,  # Add od_premium
            policy.tp_premium,  # Add tp_premium
        ])
    
    # Create a DataFrame
    columns = [
        'Insurer Name', 'Vehicle Number', 'Holder Name', 'Policy Number', 
        'Policy Issue Date', 'Policy Expiry Date', 'Policy Period', 
        'Policy Premium', 'Total Premium', 'Payment Status', 'Policy Type', 
        'Vehicle Type', 'Vehicle Make', 'Vehicle Model', 'Vehicle Gross Weight', 
        'Vehicle Manufacture Date', 'GST', 'OD Premium', 'TP Premium'
    ]
    df = pd.DataFrame(policy_data, columns=columns)

    # Generate the response for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=policy_data.xlsx'
    
    # Write to Excel using pandas ExcelWriter
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Policies')
    
    # Optional: Flash a success message
    messages.success(request, "The policy data has been successfully exported.")
    
    return response

def commission_report(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request, "Please Login First")
        return redirect('login')
    
    # Get filter values from GET parameters
    policy_no = request.GET.get("policy_no", None)
    insurer_name = request.GET.get("insurer_name", None)
    per_page = request.GET.get("per_page", 10)  # Default: 10 records per page

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    # Get logged-in user ID
    # user_id = request.user.id

    # Start with a base queryset
    policies = PolicyDocument.objects.filter(status=1)

    # Apply user role filter
    # if user_id == 2:  
    #     policies = policies.filter(rm_id=user_id)
    id  = request.user.id
    # Fetch policies
    role_id = Users.objects.filter(id=id).values_list('role_id', flat=True).first()
    if role_id != 1:
        policies = PolicyDocument.objects.filter(status=6,rm_id=id).exclude(rm_id__isnull=True).all().order_by('-id')
    else:
        policies = PolicyDocument.objects.filter(status=6).exclude(rm_id__isnull=True).all().order_by('-id')

    # Apply filters only if values are provided
    if policy_no:
        policies = policies.filter(policy_number__icontains=policy_no)
    if insurer_name:
        policies = policies.filter(insurance_provider__icontains=insurer_name)

    # Order by latest first
    policies = policies.order_by('-id')

    # Apply pagination
    paginator = Paginator(policies, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Process policy data with commission calculations
    policy_data = []
    for policy in page_obj:  # Iterate only over paginated data
        # Convert values safely
        od_premium = float(policy.od_premium.replace(',', '')) if policy.od_premium else 0.0
        tp_premium = float(policy.tp_premium.replace(',', '')) if policy.tp_premium else 0.0
        net_premium = float(policy.policy_premium.replace(',', '')) if policy.policy_premium else 0.0

        commission = policy.commission()
        if commission:
            od_percentage = float(policy.od_percent) if policy.od_percent else 0
            tp_percentage = float(policy.tp_percent	) if policy.tp_percent	 else 0
            net_percentage = float(policy.net_percent) if policy.net_percent else 0
        else:
            od_percentage = 0
            tp_percentage = 0
            net_percentage = 0

        # Calculate commission amounts
        od_commission_amount = (od_premium * od_percentage) / 100
        tp_commission_amount = (tp_premium * tp_percentage) / 100
        net_commission_amount = (net_premium * net_percentage) / 100

        policy_data.append({
            'policy': policy,
            'od_commission_amount': od_commission_amount,
            'tp_commission_amount': tp_commission_amount,
            'net_commission_amount': net_commission_amount
        })

    return render(request, 'commission_report.html', {
        'policy_data': policy_data,
        'page_obj': page_obj  # Pass paginated object to template
    })

def sales_manager_business_report(request):
    # Get filter values from GET parameters
    policy_no = request.GET.get("policy_no", None)
    insurer_name = request.GET.get("insurer_name", None)
    per_page = request.GET.get("per_page", 10)  # Default: 10 records per page

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    # Get logged-in user ID
    # user_id = request.user.id

    # Start with a base queryset
    policies = PolicyDocument.objects.filter(status=1)

    # Apply user role filter
    # if user_id == 2:  
    #     policies = policies.filter(rm_id=user_id)
    id  = request.user.id
    # Fetch policies
    role_id = Users.objects.filter(id=id).values_list('role_id', flat=True).first()
    if role_id == 2:
        policies = PolicyDocument.objects.filter(status=1,rm_id=id).exclude(rm_id__isnull=True).all().order_by('-id')
    else:
        policies = PolicyDocument.objects.filter(status=1).exclude(rm_id__isnull=True).all().order_by('-id')

    # Apply filters only if values are provided
    if policy_no:
        policies = policies.filter(policy_number__icontains=policy_no)
    if insurer_name:
        policies = policies.filter(insurance_provider__icontains=insurer_name)

    # Order by latest first
    policies = policies.order_by('-id')

    # Apply pagination
    paginator = Paginator(policies, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Process policy data with commission calculations
    policy_data = []
    for policy in page_obj:  # Iterate only over paginated data
        # Convert values safely
        od_premium = float(policy.od_premium.replace(',', '')) if policy.od_premium else 0.0
        tp_premium = float(policy.tp_premium.replace(',', '')) if policy.tp_premium else 0.0
        net_premium = float(policy.policy_premium.replace(',', '')) if policy.policy_premium else 0.0

        commission = policy.commission()
        if commission:
            od_percentage = float(policy.od_percent) if policy.od_percent else 0
            tp_percentage = float(policy.tp_percent	) if policy.tp_percent	 else 0
            net_percentage = float(policy.net_percent) if policy.net_percent else 0
        else:
            od_percentage = 0
            tp_percentage = 0
            net_percentage = 0

        # Calculate commission amounts
        od_commission_amount = (od_premium * od_percentage) / 100
        tp_commission_amount = (tp_premium * tp_percentage) / 100
        net_commission_amount = (net_premium * net_percentage) / 100

        policy_data.append({
            'policy': policy,
            'od_commission_amount': od_commission_amount,
            'tp_commission_amount': tp_commission_amount,
            'net_commission_amount': net_commission_amount
        })

    return render(request, 'reports/sales-manager-business-report.html', {
        'policy_data': policy_data,
        'page_obj': page_obj  # Pass paginated object to template
    })

def agent_business_report(request):
    # Get filter values from GET parameters
    policy_no = request.GET.get("policy_no", None)
    insurer_name = request.GET.get("insurer_name", None)
    per_page = request.GET.get("per_page", 10)  # Default: 10 records per page

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    # Get logged-in user ID
    # user_id = request.user.id

    # Start with a base queryset
    policies = PolicyDocument.objects.filter(status=1)

    # Apply user role filter
    # if user_id == 2:  
    #     policies = policies.filter(rm_id=user_id)
    id  = request.user.id
    # Fetch policies
    role_id = Users.objects.filter(id=id).values_list('role_id', flat=True).first()
    if role_id == 2:
        policies = PolicyDocument.objects.filter(status=1,rm_id=id).exclude(rm_id__isnull=True).all().order_by('-id')
    else:
        policies = PolicyDocument.objects.filter(status=1).exclude(rm_id__isnull=True).all().order_by('-id')

    # Apply filters only if values are provided
    if policy_no:
        policies = policies.filter(policy_number__icontains=policy_no)
    if insurer_name:
        policies = policies.filter(insurance_provider__icontains=insurer_name)

    # Order by latest first
    policies = policies.order_by('-id')

    # Apply pagination
    paginator = Paginator(policies, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Process policy data with commission calculations
    policy_data = []
    for policy in page_obj:  # Iterate only over paginated data
        # Convert values safely
        od_premium = float(policy.od_premium.replace(',', '')) if policy.od_premium else 0.0
        tp_premium = float(policy.tp_premium.replace(',', '')) if policy.tp_premium else 0.0
        net_premium = float(policy.policy_premium.replace(',', '')) if policy.policy_premium else 0.0

        commission = policy.commission()
        if commission:
            od_percentage = float(policy.od_percent) if policy.od_percent else 0
            tp_percentage = float(policy.tp_percent	) if policy.tp_percent	 else 0
            net_percentage = float(policy.net_percent) if policy.net_percent else 0
        else:
            od_percentage = 0
            tp_percentage = 0
            net_percentage = 0

        # Calculate commission amounts
        od_commission_amount = (od_premium * od_percentage) / 100
        tp_commission_amount = (tp_premium * tp_percentage) / 100
        net_commission_amount = (net_premium * net_percentage) / 100

        policy_data.append({
            'policy': policy,
            'od_commission_amount': od_commission_amount,
            'tp_commission_amount': tp_commission_amount,
            'net_commission_amount': net_commission_amount
        })

    return render(request, 'reports/agent-business-report.html', {
        'policy_data': policy_data,
        'page_obj': page_obj  # Pass paginated object to template
    })

def franchisees_business_report(request):
    # Get filter values from GET parameters
    policy_no = request.GET.get("policy_no", None)
    insurer_name = request.GET.get("insurer_name", None)
    per_page = request.GET.get("per_page", 10)  # Default: 10 records per page

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    # Get logged-in user ID
    # user_id = request.user.id

    # Start with a base queryset
    policies = PolicyDocument.objects.filter(status=1)

    # Apply user role filter
    # if user_id == 2:  
    #     policies = policies.filter(rm_id=user_id)
    id  = request.user.id
    # Fetch policies
    role_id = Users.objects.filter(id=id).values_list('role_id', flat=True).first()
    if role_id == 2:
        policies = PolicyDocument.objects.filter(status=1,rm_id=id).exclude(rm_id__isnull=True).all().order_by('-id')
    else:
        policies = PolicyDocument.objects.filter(status=1).exclude(rm_id__isnull=True).all().order_by('-id')

    # Apply filters only if values are provided
    if policy_no:
        policies = policies.filter(policy_number__icontains=policy_no)
    if insurer_name:
        policies = policies.filter(insurance_provider__icontains=insurer_name)

    # Order by latest first
    policies = policies.order_by('-id')

    # Apply pagination
    paginator = Paginator(policies, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Process policy data with commission calculations
    policy_data = []
    for policy in page_obj:  # Iterate only over paginated data
        # Convert values safely
        od_premium = float(policy.od_premium.replace(',', '')) if policy.od_premium else 0.0
        tp_premium = float(policy.tp_premium.replace(',', '')) if policy.tp_premium else 0.0
        net_premium = float(policy.policy_premium.replace(',', '')) if policy.policy_premium else 0.0

        commission = policy.commission()
        if commission:
            od_percentage = float(policy.od_percent) if policy.od_percent else 0
            tp_percentage = float(policy.tp_percent	) if policy.tp_percent	 else 0
            net_percentage = float(policy.net_percent) if policy.net_percent else 0
        else:
            od_percentage = 0
            tp_percentage = 0
            net_percentage = 0

        # Calculate commission amounts
        od_commission_amount = (od_premium * od_percentage) / 100
        tp_commission_amount = (tp_premium * tp_percentage) / 100
        net_commission_amount = (net_premium * net_percentage) / 100

        policy_data.append({
            'policy': policy,
            'od_commission_amount': od_commission_amount,
            'tp_commission_amount': tp_commission_amount,
            'net_commission_amount': net_commission_amount
        })

    return render(request, 'reports/franchisees-business-report.html', {
        'policy_data': policy_data,
        'page_obj': page_obj  # Pass paginated object to template
    })

def insurer_business_report(request):
    # Get filter values from GET parameters
    policy_no = request.GET.get("policy_no", None)
    insurer_name = request.GET.get("insurer_name", None)
    per_page = request.GET.get("per_page", 10)  # Default: 10 records per page

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    # Get logged-in user ID
    # user_id = request.user.id

    # Start with a base queryset
    policies = PolicyDocument.objects.filter(status=1)

    # Apply user role filter
    # if user_id == 2:  
    #     policies = policies.filter(rm_id=user_id)
    id  = request.user.id
    # Fetch policies
    role_id = Users.objects.filter(id=id).values_list('role_id', flat=True).first()
    if role_id == 2:
        policies = PolicyDocument.objects.filter(status=1,rm_id=id).exclude(rm_id__isnull=True).all().order_by('-id')
    else:
        policies = PolicyDocument.objects.filter(status=1).exclude(rm_id__isnull=True).all().order_by('-id')

    # Apply filters only if values are provided
    if policy_no:
        policies = policies.filter(policy_number__icontains=policy_no)
    if insurer_name:
        policies = policies.filter(insurance_provider__icontains=insurer_name)

    # Order by latest first
    policies = policies.order_by('-id')

    # Apply pagination
    paginator = Paginator(policies, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Process policy data with commission calculations
    policy_data = []
    for policy in page_obj:  # Iterate only over paginated data
        # Convert values safely
        od_premium = float(policy.od_premium.replace(',', '')) if policy.od_premium else 0.0
        tp_premium = float(policy.tp_premium.replace(',', '')) if policy.tp_premium else 0.0
        net_premium = float(policy.policy_premium.replace(',', '')) if policy.policy_premium else 0.0

        commission = policy.commission()
        if commission:
            od_percentage = float(policy.od_percent) if policy.od_percent else 0
            tp_percentage = float(policy.tp_percent	) if policy.tp_percent	 else 0
            net_percentage = float(policy.net_percent) if policy.net_percent else 0
        else:
            od_percentage = 0
            tp_percentage = 0
            net_percentage = 0

        # Calculate commission amounts
        od_commission_amount = (od_premium * od_percentage) / 100
        tp_commission_amount = (tp_premium * tp_percentage) / 100
        net_commission_amount = (net_premium * net_percentage) / 100

        policy_data.append({
            'policy': policy,
            'od_commission_amount': od_commission_amount,
            'tp_commission_amount': tp_commission_amount,
            'net_commission_amount': net_commission_amount
        })

    return render(request, 'reports/insurer-business-report.html', {
        'policy_data': policy_data,
        'page_obj': page_obj  # Pass paginated object to template
    })


# def download_policy_data(request):
#     # Create an in-memory Excel workbook and worksheet
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Policy Data"

#     # Define headers
#     headers = [
#         "Policy Month", "Agent Name", "SM Name", "Franchise Name", "Insurer Name", "S.P. Name",
#         "Issue Date", "Risk Start Date", "Payment Status", "Insurance Company", "Policy Type",
#         "Policy No", "Insured Name", "Vehicle Type", "Vehicle Make/Model", "Gross Weight", 
#         "Reg. No.", "MFG Year", "Sum Insured", "Gross Prem.", "GST", "Net Prem.", "OD Prem.", 
#         "TP Prem.", "Agent Comm.% OD", "Agent OD Amount", "Agent TP Comm", "Agent TP Amount", 
#         "Agent Comm.% Net", "Agent Net Amt", "Agent Bonus", "Agent Total Comm.", 
#         "Franchise Comm.% OD", "Franchise OD Amount", "Franchise TP Comm", "Franchise TP Amount", 
#         "Franchise Comm.% Net", "Franchise Net Amt", "Franchise Bonus", "Franchise Total Comm.", 
#         "Insurer Comm.% OD", "Insurer OD Amount", "Insurer TP Comm", "Insurer TP Amount", 
#         "Insurer Comm.% Net", "Insurer Net Amt", "Insurer Bonus", "Insurer Total Comm.", 
#         "Profit/Loss", "TDS %", "TDS Amount", "Net Profit"
#     ]

#     # Apply styling to header row (Blue background, White text)
#     header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
#     header_font = Font(bold=True, color="FFFFFF")

#     for col_num, header in enumerate(headers, 1):
#         cell = ws.cell(row=1, column=col_num, value=header)
#         cell.fill = header_fill
#         cell.font = header_font

#     # Default values for missing database fields
#     default_values = ["-"] * len(headers)

#     # Fetch policy data
#     policies = PolicyDocument.objects.filter(status=1).all().order_by('-id')

#     for policy in policies:
#         issue_date = policy.policy_start_date.strftime("%m-%d-%Y") if policy.policy_start_date else default_values[6]
#         issue_month = policy.policy_start_date.strftime("%b-%Y") if policy.policy_start_date else default_values[6]
#         risk_start_date = policy.start_date if policy.start_date else ""

#         # Convert string values to floats safely (handling None and removing commas)
#         def safe_float(value):
#             return float(value.replace(',', '')) if value else 0.0

#         od_premium = safe_float(policy.od_premium)
#         tp_premium = safe_float(policy.tp_premium)
#         net_premium = safe_float(policy.policy_total_premium)
#         policy_premium = safe_float(policy.policy_premium)
#         gst = safe_float(policy.gst)
#         sum_insured = safe_float(policy.sum_insured)

#         make_and_model = f"{policy.vehicle_make or '-'} / {policy.vehicle_model or '-'}"

#         # Fetch user role
#         admin_id = policy.rm_id  
#         role_id = Users.objects.filter(id=admin_id).values_list('role_id', flat=True).first()

#         # Initialize commission variables
#         od_percentage = tp_percentage = net_percentage = 0
#         od_commission_amount = tp_commission_amount = net_commission_amount = total_commission = 0.0

#         # Fetch commission data
#         commission = policy.commission()
#         if commission:
#             od_percentage = safe_float(commission.od_percentage)
#             tp_percentage = safe_float(commission.tp_percentage)
#             net_percentage = safe_float(commission.net_percentage)
#         # Calculate commissions
#             od_commission_amount = (od_premium * od_percentage) / 100
#             tp_commission_amount = (tp_premium * tp_percentage) / 100
#             net_commission_amount = (net_premium * net_percentage) / 100
#             total_commission = od_commission_amount + tp_commission_amount + net_commission_amount

#            # Insurer commission calculations
#             insurer_total_commission = 0.0
#             commission_broker = Commission.objects.filter(member_id=1).first()
#             insurer_od_percent = safe_float(commission_broker.od_percentage) if commission_broker else 0
#             insurer_tp_percent = safe_float(commission_broker.tp_percentage) if commission_broker else 0
#             insurer_net_percent = safe_float(commission_broker.net_percentage) if commission_broker else 0

#             od_commission_amount = (od_premium * insurer_od_percent) / 100
#             tp_commission_amount = (tp_premium * insurer_tp_percent) / 100
#             net_commission_amount = (net_premium * insurer_net_percent) / 100

#             insurer_total_commission = od_commission_amount + tp_commission_amount + net_commission_amount
#             profit_loss = insurer_total_commission - total_commission
            
#         # Row data for Excel
#         row_data = [
#             issue_month, policy.rm_name or default_values[1], default_values[2], default_values[3], 
#             settings.INSURER_NAME or default_values[4], default_values[5], issue_date, risk_start_date, 
#             'Confirmed', policy.insurance_provider or default_values[9], policy.policy_type or default_values[10], 
#             policy.policy_number or default_values[11], policy.holder_name or default_values[12], 
#             policy.vehicle_type or default_values[13], make_and_model, policy.vehicle_gross_weight or default_values[15], 
#             policy.vehicle_number or default_values[16], policy.vehicle_manuf_date or default_values[17], 
#             sum_insured, policy.policy_total_premium or default_values[19], gst, policy_premium, od_premium, tp_premium,
#             od_percentage if role_id != 1 else default_values[24], od_commission_amount if role_id != 1 else default_values[25],
#             tp_percentage if role_id != 1 else default_values[26], tp_commission_amount if role_id != 1 else default_values[27],
#             net_percentage if role_id != 1 else default_values[28], net_commission_amount if role_id != 1 else default_values[29],
#             0, total_commission if role_id != 1 else 0, None, None, None, None, None, None, None, None,
#             od_percentage if role_id in [1, 2] else None, od_commission_amount if role_id in [1, 2] else None,
#             tp_percentage if role_id in [1, 2] else None, tp_commission_amount if role_id in [1, 2] else None,
#             net_percentage if role_id in [1, 2] else None, net_commission_amount if role_id in [1, 2] else None,
#             0, total_commission, profit_loss, 0, 0, profit_loss
#         ]
#         ws.append(row_data)

#     # Create HTTP response for downloading
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="policy_data.xlsx"'
#     wb.save(response)

#     return response

def download_policy_data(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Policy Data"
    user_id = request.user.id

    # Define headers based on role
    full_headers = [
        "Policy Month", "Agent Name", "SM Name", "Franchise Name", "Insurer Name", "S.P. Name",
        "Issue Date", "Risk Start Date", "Payment Status", "Insurance Company", "Policy Type",
        "Policy No", "Insured Name", "Vehicle Type", "Vehicle Make/Model", "Gross Weight", 
        "Reg. No.", "MFG Year", "Sum Insured", "Gross Prem.", "GST", "Net Prem.", "OD Prem.", 
        "TP Prem.", "Agent Comm.% OD", "Agent OD Amount", "Agent TP Comm", "Agent TP Amount", 
        "Agent Comm.% Net", "Agent Net Amt", "Agent Bonus", "Agent Total Comm.", 
        "Franchise Comm.% OD", "Franchise OD Amount", "Franchise TP Comm", "Franchise Agent TP Amount", 
        "Franchise Agent Comm.% Net", "Franchise Agent Net Amt", "Franchise Bonus", "Franchise Total Comm.", 
        "Insurer Comm.% OD", "Insurer OD Amount", "Insurer TP Comm", "Insurer TP Amount", 
        "Insurer Comm.% Net", "Insurer Net Amt", "Insurer Bonus", "Insurer Total Comm.", 
        "Profit/Loss", "TDS %", "TDS Amount", "Net Profit"
    ]
    
    limited_headers = full_headers[:32]  # Show only up to "Agent Total Comm." for role ID 2
    headers = full_headers if user_id == 1 else limited_headers

    # Apply styling to the header row
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
    id  = request.user.id
    # Fetch policies
    role_id = Users.objects.filter(id=id).values_list('role_id', flat=True).first()
    if role_id != 1:
        policies = PolicyDocument.objects.filter(status=6,rm_id=id).all().order_by('-id')
    else:
        policies = PolicyDocument.objects.filter(status=6).all().order_by('-id')

    for policy in policies:
        issue_month = "-"
        issue_date = "-"
        if policy.policy_start_date:
            try:
                start_date_obj = policy.policy_start_date if isinstance(policy.policy_start_date, datetime) else datetime.strptime(policy.policy_start_date, "%Y-%m-%d")
                issue_month = start_date_obj.strftime("%b-%Y")
                issue_date = start_date_obj.strftime("%m-%d-%Y")
            except Exception as e:
                print(f"Date conversion error: {e}")
        risk_start_date = policy.start_date or "-"
        
        od_premium = float(policy.od_premium.replace(',', '')) if policy.od_premium else 0.0  
        tp_premium = float(policy.tp_premium.replace(',', '')) if policy.tp_premium else 0.0  
        net_premium = float(policy.policy_premium.replace(',', '')) if policy.policy_premium else 0.0
        
        admin_id = policy.rm_id  
        role_id = Users.objects.filter(id=admin_id).values_list('role_id', flat=True).first()
        
        od_percentage = float(policy.od_percent) if policy.od_percent else 0
        tp_percentage = float(policy.tp_percent) if policy.tp_percent else 0
        net_percentage = float(policy.net_percent) if policy.net_percent else 0
        
        if role_id == 2:  # Agent Role
            od_commission_amount = (od_premium * od_percentage) / 100
            tp_commission_amount = (tp_premium * tp_percentage) / 100
            net_commission_amount = (net_premium * net_percentage) / 100
            agent_total_commission = od_commission_amount + tp_commission_amount + net_commission_amount
        else:
            od_commission_amount = tp_commission_amount = net_commission_amount = agent_total_commission = 0

        # Broker commission calculations
        commission_broker = Commission.objects.filter(member_id=1).first()
        insurer_od_percent = float(policy.insurer_od_commission) if policy.insurer_od_commission else 0
        insurer_tp_percent = float(policy.insurer_tp_commission) if policy.insurer_tp_commission else 0
        insurer_net_percent = float(policy.insurer_net_commission) if policy.insurer_net_commission else 0 
        
        insurer_od_commission = (od_premium * insurer_od_percent) / 100
        insurer_tp_commission = (tp_premium * insurer_tp_percent) / 100
        insurer_net_commission = (net_premium * insurer_net_percent) / 100
        insurer_total_commission = insurer_od_commission + insurer_tp_commission + insurer_net_commission

        profit_loss = insurer_total_commission - agent_total_commission if role_id == 2 else insurer_total_commission

        row_data = [
            issue_month, policy.rm_name or "-", "-", "-", settings.INSURER_NAME or "-", "-",
            issue_date, risk_start_date, 'Confirmed', policy.insurance_provider or "-", policy.policy_type or "-",
            policy.policy_number or "-", policy.holder_name or "-", policy.vehicle_type or "-",
            f"{policy.vehicle_make}/{policy.vehicle_model}" if policy.vehicle_make and policy.vehicle_model else "-", 
            policy.vehicle_gross_weight or "-", policy.vehicle_number or "-", policy.vehicle_manuf_date or "-",
            policy.sum_insured or "-", policy.policy_total_premium or "-", policy.gst or "-", policy.policy_premium or "-",
            policy.od_premium or "-", policy.tp_premium or "-",
            od_percentage if role_id == 2 else "-",
            od_commission_amount if role_id == 2 else "-",
            tp_percentage if role_id == 2 else "-",
            tp_commission_amount if role_id == 2 else "-",
            net_percentage if role_id == 2 else "-",
            net_commission_amount if role_id == 2 else "-",
            "-", agent_total_commission if role_id == 2 else "-",
        ]

        if user_id == 1:  # Append extra fields only for role ID 1
            row_data.extend([
                None,None,None,None,None,None,None,None,
                insurer_od_percent, insurer_od_commission, insurer_tp_percent, insurer_tp_commission,
                insurer_net_percent, insurer_net_commission, "-", insurer_total_commission,
                profit_loss, "-", "-", profit_loss
            ])
        ws.append(row_data)

    # Generate Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="policy_data.xlsx"'
    wb.save(response)
    return response

def export_commission_data(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request, "Please Login First")
        return redirect('login')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Policy Data"
    user_id = request.user.id

    # Define headers based on role
    full_headers = [
        "Policy Month", "Agent Name", "SM Name", "Franchise Name", "Insurer Name", "S.P. Name",
        "Issue Date", "Risk Start Date", "Payment Status", "Insurance Company", "Policy Type",
        "Policy No", "Insured Name", "Vehicle Type", "Vehicle Make/Model", "Gross Weight", 
        "Reg. No.", "MFG Year", "Sum Insured", "Gross Prem.", "GST", "Net Prem.", "OD Prem.", 
        "TP Prem.", "Agent Comm.% OD", "Agent OD Amount", "Agent TP Comm", "Agent TP Amount", 
        "Agent Comm.% Net", "Agent Net Amt", "Agent Bonus", "Agent Total Comm.", 
        "Franchise Comm.% OD", "Franchise OD Amount", "Franchise TP Comm", "Franchise Agent TP Amount", 
        "Franchise Agent Comm.% Net", "Franchise Agent Net Amt", "Franchise Bonus", "Franchise Total Comm.", 
        "Insurer Comm.% OD", "Insurer OD Amount", "Insurer TP Comm", "Insurer TP Amount", 
        "Insurer Comm.% Net", "Insurer Net Amt", "Insurer Bonus", "Insurer Total Comm.", 
        "Profit/Loss", "TDS %", "TDS Amount", "Net Profit"
    ]
    
    limited_headers = full_headers[:32]
    headers = full_headers if user_id == 1 else limited_headers

    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
    id  = request.user.id
    
    # Fetch policies
    role_id = Users.objects.filter(id=id).values_list('role_id', flat=True).first()
    if role_id != 1:
        policies = PolicyDocument.objects.filter(status=6,rm_id=id).all().order_by('-id')
    else:
        policies = PolicyDocument.objects.filter(status=6).all().order_by('-id')

    for policy in policies:
        issue_month = "-"
        issue_date = "-"
        if policy.policy_start_date:
            try:
                start_date_obj = policy.policy_start_date if isinstance(policy.policy_start_date, datetime) else datetime.strptime(policy.policy_start_date, "%Y-%m-%d")
                issue_month = start_date_obj.strftime("%b-%Y")
                issue_date = start_date_obj.strftime("%m-%d-%Y")
            except Exception as e:
                print(f"Date conversion error: {e}")
        risk_start_date = policy.start_date or "-"
        
        agent_payment_info = AgentPaymentDetails.objects.filter(policy_number=policy.policy_number).first()
        insurer_payment_info = InsurerPaymentDetails.objects.filter(policy_number=policy.policy_number).first()
        
        admin_id = policy.rm_id  
        role_id = Users.objects.filter(id=admin_id).values_list('role_id', flat=True).first()
        
        insurer_total_comm_amt = float(getattr(insurer_payment_info, 'insurer_total_comm_amount', 0) or 0)
        agent_total_comm_amt = float(getattr(agent_payment_info, 'agent_total_comm_amount', 0) or 0)
        profit_loss = insurer_total_comm_amt - agent_total_comm_amt
                
        row_data = [
            issue_month,
            policy.rm_name or "-",
            "-", "-", "-",
            settings.INSURER_NAME or "-",
            issue_date,
            risk_start_date,
            'Confirmed',
            policy.insurance_provider or "-",
            policy.policy_type or "-",
            policy.policy_number or "-",
            policy.holder_name or "-",
            policy.vehicle_type or "-",
            f"{policy.vehicle_make}/{policy.vehicle_model}" if policy.vehicle_make and policy.vehicle_model else "-",
            policy.vehicle_gross_weight or "-",
            policy.vehicle_number or "-",
            policy.vehicle_manuf_date or "-",
            policy.sum_insured or "-",
            policy.policy_total_premium or "-",
            policy.gst or "-",
            policy.policy_premium or "-",
            policy.od_premium or "-",
            policy.tp_premium or "-",
            agent_payment_info.agent_od_comm if agent_payment_info else "-",
            agent_payment_info.agent_od_amount if agent_payment_info else "-",
            agent_payment_info.agent_tp_comm if agent_payment_info else "-",
            agent_payment_info.agent_tp_amount if agent_payment_info else "-",
            agent_payment_info.agent_net_comm if agent_payment_info else "-",
            agent_payment_info.agent_net_amount if agent_payment_info else "-",
            agent_payment_info.agent_incentive_amount if agent_payment_info else "-",
            agent_payment_info.agent_total_comm_amount if agent_payment_info else "-",
        ]

        # Add insurer-specific fields if user is admin (user_id == 1)
        if user_id == 1:
            row_data.extend([
                None, None, None, None, None, None, None, None,
                insurer_payment_info.insurer_od_comm if insurer_payment_info and insurer_payment_info.insurer_od_comm else "-",
                insurer_payment_info.insurer_od_amount if insurer_payment_info and insurer_payment_info.insurer_od_amount else "-",
                insurer_payment_info.insurer_tp_comm if insurer_payment_info and insurer_payment_info.insurer_tp_comm else "-",
                insurer_payment_info.insurer_tp_amount if insurer_payment_info and insurer_payment_info.insurer_tp_amount else "-",
                insurer_payment_info.insurer_net_comm if insurer_payment_info and insurer_payment_info.insurer_net_comm else "-",
                insurer_payment_info.insurer_net_amount if insurer_payment_info and insurer_payment_info.insurer_net_amount else "-",
                insurer_payment_info.insurer_incentive_amount if insurer_payment_info and insurer_payment_info.insurer_incentive_amount else "-",
                insurer_payment_info.insurer_total_comm_amount if insurer_payment_info and insurer_payment_info.insurer_total_comm_amount else "-",
                profit_loss if profit_loss else "-",
                insurer_payment_info.insurer_tds if insurer_payment_info and insurer_payment_info.insurer_tds else "-",
                insurer_payment_info.insurer_tds_amount if insurer_payment_info and insurer_payment_info.insurer_tds_amount else "-",
                insurer_payment_info.insurer_balance_amount if insurer_payment_info and insurer_payment_info.insurer_balance_amount else "-"
            ])
        ws.append(row_data)

    # Generate Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="policy_data.xlsx"'
    wb.save(response)
    return response

def export_commission_data_v1(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request, "Please Login First")
        return redirect('login')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Policy Data"
    user_id = request.user.id

    # Define headers based on role
    full_headers = [
        "Policy Month", "Agent Name", "SM Name", "Franchise Name", "Insurer Name", "S.P. Name",
        "Issue Date", "Risk Start Date", "Payment Status", "Insurance Company", "Policy Type",
        "Policy No", "Insured Name", "Vehicle Type", "Vehicle Make/Model", "Gross Weight", 
        "Reg. No.", "MFG Year", "Sum Insured", "Gross Prem.", "GST", "Net Prem.", "OD Prem.", 
        "TP Prem.", "Agent Comm.% OD", "Agent OD Amount", "Agent TP Comm", "Agent TP Amount", 
        "Agent Comm.% Net", "Agent Net Amt", "Agent Bonus", "Agent Total Comm.", 
        "Franchise Comm.% OD", "Franchise OD Amount", "Franchise TP Comm", "Franchise Agent TP Amount", 
        "Franchise Agent Comm.% Net", "Franchise Agent Net Amt", "Franchise Bonus", "Franchise Total Comm.", 
        "Insurer Comm.% OD", "Insurer OD Amount", "Insurer TP Comm", "Insurer TP Amount", 
        "Insurer Comm.% Net", "Insurer Net Amt", "Insurer Bonus", "Insurer Total Comm.", 
        "Profit/Loss", "TDS %", "TDS Amount", "Net Profit"
    ]
    
    limited_headers = full_headers[:32]
    headers = full_headers if user_id == 1 else limited_headers

    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
    id  = request.user.id
    
     # Get filter values from GET parameters
    policy_no = request.GET.get("policy_no", None)
    insurer_name = request.GET.get("insurer_name", None)
    service_provider = request.GET.get("service_provider", None)
    insurance_company = request.GET.get("insurance_company", None)
    policy_type = request.GET.get("policy_type", None)
    vehicle_type = request.GET.get("vehicle_type", None)
    referral_name = request.GET.get("referral_name", None)
    vehicle_reg_no = request.GET.get("vehicle_reg_no", None)
    policy_start_date = request.GET.get("policy_start_date", None)
    policy_end_date = request.GET.get("policy_end_date", None)
    per_page = request.GET.get("per_page", 20)  # Default: 20 records per page
    page_number = request.GET.get('page',1)
    
    
    user_id  = request.user.id
    role_id = request.user.role_id
    
    if role_id != 1:
        policies = PolicyDocument.objects.filter(status=6,rm_id=user_id).exclude(rm_id__isnull=True).all()
    else:
        policies = PolicyDocument.objects.filter(status=6).exclude(rm_id__isnull=True).all()

    if policy_no:
        policies = policies.filter(policy_number__icontains=policy_no)
    
    if insurer_name:
        policies = policies.filter(policy_info__insurance_company__icontains=insurer_name)  # <-- yeh policy_info vali table se data check krega
    
    if service_provider:
        policies = policies.filter(policy_info__service_provider__icontains=service_provider)  # <-- yeh policy_info vali table se data check krega
        
    if insurance_company:
        policies = policies.filter(policy_info__insurance_company__icontains=insurance_company)  # <-- yeh policy_info vali table se data check krega
    
    if policy_type:
        policies = policies.filter(policy_info__policy_type__icontains=policy_type)  # <-- yeh policy_info vali table se data check krega
        
    if vehicle_type:
        policies = policies.filter(policy_vehicle_info__vehicle_type__icontains=vehicle_type)  # <-- yeh policy_vehicle_info vali table se data check krega
    
    if referral_name:
        policies = policies.filter(policy_agent_info__referral__name__icontains=referral_name)  # <-- yeh agent_payment_details vali table se data check krega

    if vehicle_reg_no:
        policies = policies.filter(policy_vehicle_info__registration_number__icontains=vehicle_reg_no)  # <-- yeh policy_vehicle_info vali table se data check krega

    if policy_start_date:
        policies = policies.filter(policy_info__policy_start_date__icontains=policy_start_date)  # <-- yeh policy_info vali table se data check krega

    if policy_end_date:
        policies = policies.filter(policy_info__policy_end_date__icontains=policy_end_date)  # <-- yeh policy_info vali table se data check krega


    policies = policies.order_by('-id')
    
    for policy in policies:
        issue_month = "-"
        issue_date = "-"
        if policy.policy_start_date:
            try:
                start_date_obj = policy.policy_start_date if isinstance(policy.policy_start_date, datetime) else datetime.strptime(policy.policy_start_date, "%Y-%m-%d")
                issue_month = start_date_obj.strftime("%b-%Y")
                issue_date = start_date_obj.strftime("%m-%d-%Y")
            except Exception as e:
                print(f"Date conversion error: {e}")
        risk_start_date = policy.start_date or "-"
        
        agent_payment_info = AgentPaymentDetails.objects.filter(policy_number=policy.policy_number).first()
        insurer_payment_info = InsurerPaymentDetails.objects.filter(policy_number=policy.policy_number).first()
        
        admin_id = policy.rm_id  
        role_id = Users.objects.filter(id=admin_id).values_list('role_id', flat=True).first()
        
        insurer_total_comm_amt = float(getattr(insurer_payment_info, 'insurer_total_comm_amount', 0) or 0)
        agent_total_comm_amt = float(getattr(agent_payment_info, 'agent_total_comm_amount', 0) or 0)
        profit_loss = insurer_total_comm_amt - agent_total_comm_amt
                
        row_data = [
            issue_month,
            policy.rm_name or "-",
            "-", "-", "-",
            settings.INSURER_NAME or "-",
            issue_date,
            risk_start_date,
            'Confirmed',
            policy.insurance_provider or "-",
            policy.policy_type or "-",
            policy.policy_number or "-",
            policy.holder_name or "-",
            policy.vehicle_type or "-",
            f"{policy.vehicle_make}/{policy.vehicle_model}" if policy.vehicle_make and policy.vehicle_model else "-",
            policy.vehicle_gross_weight or "-",
            policy.vehicle_number or "-",
            policy.vehicle_manuf_date or "-",
            policy.sum_insured or "-",
            policy.policy_total_premium or "-",
            policy.gst or "-",
            policy.policy_premium or "-",
            policy.od_premium or "-",
            policy.tp_premium or "-",
            agent_payment_info.agent_od_comm if agent_payment_info else "-",
            agent_payment_info.agent_od_amount if agent_payment_info else "-",
            agent_payment_info.agent_tp_comm if agent_payment_info else "-",
            agent_payment_info.agent_tp_amount if agent_payment_info else "-",
            agent_payment_info.agent_net_comm if agent_payment_info else "-",
            agent_payment_info.agent_net_amount if agent_payment_info else "-",
            agent_payment_info.agent_incentive_amount if agent_payment_info else "-",
            agent_payment_info.agent_total_comm_amount if agent_payment_info else "-",
        ]

        # Add insurer-specific fields if user is admin (user_id == 1)
        if user_id == 1:
            row_data.extend([
                None, None, None, None, None, None, None, None,
                insurer_payment_info.insurer_od_comm if insurer_payment_info and insurer_payment_info.insurer_od_comm else "-",
                insurer_payment_info.insurer_od_amount if insurer_payment_info and insurer_payment_info.insurer_od_amount else "-",
                insurer_payment_info.insurer_tp_comm if insurer_payment_info and insurer_payment_info.insurer_tp_comm else "-",
                insurer_payment_info.insurer_tp_amount if insurer_payment_info and insurer_payment_info.insurer_tp_amount else "-",
                insurer_payment_info.insurer_net_comm if insurer_payment_info and insurer_payment_info.insurer_net_comm else "-",
                insurer_payment_info.insurer_net_amount if insurer_payment_info and insurer_payment_info.insurer_net_amount else "-",
                insurer_payment_info.insurer_incentive_amount if insurer_payment_info and insurer_payment_info.insurer_incentive_amount else "-",
                insurer_payment_info.insurer_total_comm_amount if insurer_payment_info and insurer_payment_info.insurer_total_comm_amount else "-",
                profit_loss if profit_loss else "-",
                insurer_payment_info.insurer_tds if insurer_payment_info and insurer_payment_info.insurer_tds else "-",
                insurer_payment_info.insurer_tds_amount if insurer_payment_info and insurer_payment_info.insurer_tds_amount else "-",
                insurer_payment_info.insurer_balance_amount if insurer_payment_info and insurer_payment_info.insurer_balance_amount else "-"
            ])
        ws.append(row_data)

    # Generate Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="policy_data.xlsx"'
    wb.save(response)
    return response



